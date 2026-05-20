"""
Report Service - Generate PDF and Excel reports
Handles all report generation, formatting, and export functionality
"""

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.pdfgen import canvas as pdf_canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from datetime import datetime, timedelta
import io
import traceback


class ReportService:
    """Service for generating PDF and Excel reports"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # Heading style
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#166534'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))
        
        # Subheading style
        self.styles.add(ParagraphStyle(
            name='CustomSubheading',
            parent=self.styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#0f172a'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        ))
    
    def generate_pdf_report(self, report_data, report_type='overview'):
        """
        Generate PDF report
        
        Args:
            report_data (dict): Report data containing analytics
            report_type (str): Type of report (daily, weekly, monthly, custom, overview)
        
        Returns:
            BytesIO: PDF file buffer
        """
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter,
                                   rightMargin=72, leftMargin=72,
                                   topMargin=72, bottomMargin=18)
            
            # Container for the 'Flowable' objects
            elements = []
            
            # Add title
            title_text = self._get_report_title(report_type, report_data)
            title = Paragraph(title_text, self.styles['CustomTitle'])
            elements.append(title)
            elements.append(Spacer(1, 12))
            
            # Add generation date
            date_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            date_para = Paragraph(date_text, self.styles['Normal'])
            elements.append(date_para)
            elements.append(Spacer(1, 20))
            
            # Add executive summary
            if report_data.get('include_overview', True):
                elements.extend(self._create_executive_summary(report_data))
                elements.append(Spacer(1, 20))
            
            # Add key metrics table
            if report_data.get('key_metrics'):
                elements.extend(self._create_metrics_table(report_data['key_metrics']))
                elements.append(Spacer(1, 20))
            
            # Add analytics table
            if report_data.get('analytics_data'):
                elements.extend(self._create_analytics_table(report_data['analytics_data']))
                elements.append(Spacer(1, 20))
            
            # Add department performance
            if report_data.get('include_departments', True) and report_data.get('dept_performance'):
                elements.extend(self._create_department_section(report_data['dept_performance']))
                elements.append(Spacer(1, 20))
            
            # Add reward analysis
            if report_data.get('include_rewards', False) and report_data.get('reward_categories'):
                elements.extend(self._create_rewards_section(report_data['reward_categories']))
            
            # Build PDF
            doc.build(elements)
            
            # Get PDF data
            buffer.seek(0)
            return buffer
            
        except Exception as e:
            print(f"Error generating PDF report: {e}")
            traceback.print_exc()
            raise
    
    def generate_excel_report(self, report_data, report_type='overview'):
        """
        Generate Excel report
        
        Args:
            report_data (dict): Report data containing analytics
            report_type (str): Type of report
        
        Returns:
            BytesIO: Excel file buffer
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add Overview sheet
            self._create_overview_sheet(wb, report_data, report_type)
            
            # Add Analytics sheet
            if report_data.get('analytics_data'):
                self._create_analytics_sheet(wb, report_data['analytics_data'])
            
            # Add Department Performance sheet
            if report_data.get('dept_performance'):
                self._create_department_sheet(wb, report_data['dept_performance'])
            
            # Add Rewards Analysis sheet
            if report_data.get('reward_categories'):
                self._create_rewards_sheet(wb, report_data['reward_categories'])
            
            # Add Weekly Trend sheet
            if report_data.get('weekly_trend'):
                self._create_weekly_trend_sheet(wb, report_data['weekly_trend'])
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            print(f"Error generating Excel report: {e}")
            traceback.print_exc()
            raise
    
    def _get_report_title(self, report_type, report_data):
        """Get report title based on type"""
        titles = {
            'daily': 'Daily Summary Report',
            'weekly': 'Weekly Analytics Report',
            'monthly': 'Monthly Performance Report',
            'custom': report_data.get('report_name', 'Custom Report'),
            'overview': 'System Overview Report'
        }
        return titles.get(report_type, 'Trash to Treasure Report')
    
    def _create_executive_summary(self, report_data):
        """Create executive summary section"""
        elements = []
        
        heading = Paragraph("Executive Summary", self.styles['CustomHeading'])
        elements.append(heading)
        
        key_metrics = report_data.get('key_metrics', {})
        
        summary_text = f"""
        This report provides a comprehensive overview of the Trash to Treasure recycling program.
        During the reporting period, the system recorded <b>{key_metrics.get('total_bottles', {}).get('value', 0)} bottles</b> collected,
        with <b>{key_metrics.get('points_circulated', {}).get('value', 0)} points</b> circulated among
        <b>{key_metrics.get('active_users', {}).get('value', 0)} active users</b>.
        A total of <b>{key_metrics.get('rewards_redeemed', {}).get('value', 0)} rewards</b> were redeemed.
        Machine uptime averaged <b>{key_metrics.get('machine_uptime', {}).get('value', 'N/A')}</b>.
        """
        
        summary = Paragraph(summary_text, self.styles['Normal'])
        elements.append(summary)
        
        return elements
    
    def _create_metrics_table(self, key_metrics):
        """Create key metrics comparison table"""
        elements = []
        
        heading = Paragraph("Key Performance Indicators", self.styles['CustomHeading'])
        elements.append(heading)
        
        # Prepare table data
        data = [['Metric', 'Current Value', 'Change vs Last Period']]
        
        metrics_mapping = {
            'total_bottles': 'Total Bottles Collected',
            'points_circulated': 'Points Circulated',
            'active_users': 'Active Users',
            'rewards_redeemed': 'Rewards Redeemed',
            'machine_uptime': 'Machine Uptime'
        }
        
        for key, label in metrics_mapping.items():
            metric_data = key_metrics.get(key, {})
            value = metric_data.get('value', 0)
            change = metric_data.get('change', 0)
            change_text = f"+{change}%" if change >= 0 else f"{change}%"
            data.append([label, str(value), change_text])
        
        # Create table
        table = Table(data, colWidths=[3*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ]))
        
        elements.append(table)
        return elements
    
    def _create_analytics_table(self, analytics_data):
        """Create detailed analytics table"""
        elements = []
        
        heading = Paragraph("Detailed Analytics by Period", self.styles['CustomHeading'])
        elements.append(heading)
        
        # Prepare table data
        data = [['Period', 'Bottles', 'Points', 'Rewards', 'Users', 'Top Dept', 'Uptime']]
        
        for period in analytics_data:
            data.append([
                period.get('period', 'N/A'),
                str(period.get('bottles', 0)),
                str(period.get('points', 0)),
                str(period.get('rewards', 0)),
                str(period.get('users', 0)),
                period.get('top_dept', 'N/A'),
                period.get('uptime', 'N/A')
            ])
        
        # Create table  (7 cols × 6.4" total — fits in 6.5" frame)
        table = Table(data, colWidths=[1.0*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1.3*inch, 0.75*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#166534')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(table)
        return elements
    
    def _create_department_section(self, dept_performance):
        """Create department performance section"""
        elements = []
        
        heading = Paragraph("Department Performance Analysis", self.styles['CustomHeading'])
        elements.append(heading)
        
        # Prepare table data
        data = [['Department', 'Points', 'Bottles', 'Users']]
        
        for dept in dept_performance[:10]:  # Top 10 departments
            data.append([
                dept.get('name', 'Unknown'),
                str(dept.get('points', 0)),
                str(dept.get('bottles', 0)),
                str(dept.get('users', 0))
            ])
        
        # Create table  (4 cols × 6.3" total — fits in 6.5" frame)
        table = Table(data, colWidths=[2.7*inch, 1.2*inch, 1.2*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        return elements
    
    def _create_rewards_section(self, reward_categories):
        """Create rewards analysis section"""
        elements = []
        
        heading = Paragraph("Reward Redemption Analysis", self.styles['CustomHeading'])
        elements.append(heading)
        
        # Prepare table data
        data = [['Reward Name', 'Redemptions']]
        
        for reward_name, count in reward_categories[:10]:
            data.append([reward_name, str(count)])
        
        # Create table
        table = Table(data, colWidths=[4*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightyellow),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        return elements
    
    def _create_overview_sheet(self, wb, report_data, report_type):
        """Create overview sheet in Excel"""
        ws = wb.create_sheet("Overview")
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        
        # Title
        ws['A1'] = self._get_report_title(report_type, report_data)
        ws['A1'].font = Font(size=16, bold=True, color="1E40AF")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        # Generation date
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:B2')
        
        # Key Metrics
        row = 4
        ws[f'A{row}'] = "Key Performance Indicators"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1
        
        # Headers
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        ws[f'B{row}'].fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        row += 1
        
        # Metrics data
        key_metrics = report_data.get('key_metrics', {})
        metrics_mapping = {
            'total_bottles': 'Total Bottles Collected',
            'points_circulated': 'Points Circulated',
            'active_users': 'Active Users',
            'rewards_redeemed': 'Rewards Redeemed',
            'machine_uptime': 'Machine Uptime'
        }
        
        for key, label in metrics_mapping.items():
            metric_data = key_metrics.get(key, {})
            value = metric_data.get('value', 0)
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
    
    def _create_analytics_sheet(self, wb, analytics_data):
        """Create analytics sheet in Excel"""
        ws = wb.create_sheet("Period Analytics")
        
        # Headers
        headers = ['Period', 'Bottles', 'Points', 'Rewards', 'Users', 'Top Department', 'Uptime']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, period in enumerate(analytics_data, start=2):
            ws.cell(row=row_idx, column=1, value=period.get('period', 'N/A'))
            ws.cell(row=row_idx, column=2, value=period.get('bottles', 0))
            ws.cell(row=row_idx, column=3, value=period.get('points', 0))
            ws.cell(row=row_idx, column=4, value=period.get('rewards', 0))
            ws.cell(row=row_idx, column=5, value=period.get('users', 0))
            ws.cell(row=row_idx, column=6, value=period.get('top_dept', 'N/A'))
            ws.cell(row=row_idx, column=7, value=period.get('uptime', 'N/A'))
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_department_sheet(self, wb, dept_performance):
        """Create department performance sheet in Excel"""
        ws = wb.create_sheet("Departments")
        
        # Headers
        headers = ['Department', 'Points', 'Bottles', 'Users']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, dept in enumerate(dept_performance, start=2):
            ws.cell(row=row_idx, column=1, value=dept.get('name', 'Unknown'))
            ws.cell(row=row_idx, column=2, value=dept.get('points', 0))
            ws.cell(row=row_idx, column=3, value=dept.get('bottles', 0))
            ws.cell(row=row_idx, column=4, value=dept.get('users', 0))
        
        # Add chart
        if len(dept_performance) > 0:
            chart = BarChart()
            chart.title = "Department Performance (Points)"
            chart.x_axis.title = "Department"
            chart.y_axis.title = "Points"
            
            data = Reference(ws, min_col=2, min_row=1, max_row=len(dept_performance)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(dept_performance)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "F2")
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_rewards_sheet(self, wb, reward_categories):
        """Create rewards analysis sheet in Excel"""
        ws = wb.create_sheet("Rewards")
        
        # Headers
        ws['A1'] = "Reward Name"
        ws['B1'] = "Redemptions"
        ws['A1'].font = Font(bold=True, color="FFFFFF")
        ws['B1'].font = Font(bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        ws['B1'].fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['B1'].alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, (reward_name, count) in enumerate(reward_categories, start=2):
            ws.cell(row=row_idx, column=1, value=reward_name)
            ws.cell(row=row_idx, column=2, value=count)
        
        # Add pie chart
        if len(reward_categories) > 0:
            chart = PieChart()
            chart.title = "Reward Distribution"
            
            data = Reference(ws, min_col=2, min_row=1, max_row=len(reward_categories)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(reward_categories)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "D2")
        
        # Auto-size columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _create_weekly_trend_sheet(self, wb, weekly_trend):
        """Create weekly trend sheet in Excel"""
        ws = wb.create_sheet("Weekly Trend")
        
        # Headers
        headers = ['Day', 'Bottles', 'Points', 'Users']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, day_data in enumerate(weekly_trend, start=2):
            ws.cell(row=row_idx, column=1, value=day_data.get('date', ''))
            ws.cell(row=row_idx, column=2, value=day_data.get('bottles', 0))
            ws.cell(row=row_idx, column=3, value=day_data.get('points', 0))
            ws.cell(row=row_idx, column=4, value=day_data.get('users', 0))
        
        # Add line chart
        if len(weekly_trend) > 0:
            chart = LineChart()
            chart.title = "Weekly Activity Trend"
            chart.x_axis.title = "Day"
            chart.y_axis.title = "Count"
            
            data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=len(weekly_trend)+1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(weekly_trend)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "F2")
        
        # Auto-size columns
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15
